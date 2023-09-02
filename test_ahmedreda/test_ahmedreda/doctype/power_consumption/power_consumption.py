# Copyright (c) 2023, Ahmed Reda and contributors
# For license information, please see license.txt

import frappe
# openpyxl is a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files
import openpyxl as xl
from calendar import month_name
from frappe.model.document import Document

class PowerConsumption(Document):
	def validate(self):
		self.validate_file_type()

	def validate_file_type(self):
		# Make sure the file extension is xlsx
		if self.excel_file.split(".")[1] != 'xlsx':
			frappe.throw('Invalid File Type, It must end with .xlsx')

	@frappe.whitelist()
	def get_master_data(self):
		# logger = frappe.logger("api", allow_site=True, file_count=50)
		# logger.info(f"sheet.cell(row, 2).value ={sheet.cell(2, 2).value}")
		file_name = self.excel_file.split("/")
		file_path = frappe.get_site_path('private', 'files', file_name[-1])
		workbook = xl.load_workbook(file_path)
		first_sheet = workbook[workbook.sheetnames[0]]
		return {
			#.cell(#row, #column)
			'customer_name': first_sheet.cell(2, 2).value, 
			'phone':first_sheet.cell(3, 2).value,
			'project':first_sheet.cell(4, 2).value
		}
	
	@frappe.whitelist()
	def get_calculations(self):
		file_name = self.excel_file.split("/")
		file_path = frappe.get_site_path('private', 'files', file_name[-1])
		workbook = xl.load_workbook(file_path)
		first_sheet = workbook[workbook.sheetnames[0]]
		total_of_kw = 0
		total_of_kwh =  0 
		number_of_rows  = 0
		total_of_low_tariﬀs = 0
		counter_for_low_tariffs = 0  
		total_of_high_tariﬀs = 0
		counter_for_high_tariffs = 0 
		tariﬀs_list = []
		start_row = 8
		total_rows = first_sheet.max_row - start_row
		counter = 0
		for row in range(start_row, first_sheet.max_row):
			counter +=1
			frappe.publish_progress(counter*100/(total_rows),title="Make Tariifs Calculations...",
                                description = f"""{counter} Of {total_rows} Calculate""")
			
			curr_date_time = first_sheet.cell(row, 1).value
			previous_date_time = first_sheet.cell(row, 1).value
			if row == start_row:
				previous_date_time = curr_date_time
			else:
				previous_date_time = first_sheet.cell(row - 1, 1).value

			kw = first_sheet.cell(row, 2).value
			kwh = first_sheet.cell(row, 3).value

			# Calculate  Avg  of KW and  KWH.
			if isinstance(kw,int) or isinstance(kw,float) or isinstance(kwh,int) or isinstance(kwh,float):
				number_of_rows += 1
			if isinstance(kw,int) or isinstance(kw,float):
				total_of_kw += kw
			if isinstance(kwh,int) or isinstance(kwh,float):
				total_of_kwh += kwh
			# End of Calculations.

			# Calculate the average low and average high tariﬀs for each month
			if isinstance(kwh,int) or isinstance(kwh,float):#
				# High tariﬀs
				if (curr_date_time.hour == 23 or curr_date_time.hour ==0) or (curr_date_time.hour >= 1 and curr_date_time.hour <=5):
					if curr_date_time.month == previous_date_time.month:
						total_of_high_tariﬀs += kwh
						counter_for_high_tariffs += 1
					else:
						tariﬀs_list.append({
							'month':month_name[previous_date_time.month],
							'average_high_tariﬀs': 0.3*(total_of_high_tariﬀs/counter_for_high_tariffs),
							'average_low_tariﬀs': 0.1*(total_of_low_tariﬀs/counter_for_low_tariffs),
							'year':previous_date_time.year
						})

						total_of_high_tariﬀs = kwh
						counter_for_high_tariffs = 1
						total_of_low_tariﬀs = 0
						counter_for_low_tariffs = 0
				# Low tariﬀs
				elif (curr_date_time.hour >= 6 and curr_date_time.hour <=12) or (curr_date_time.hour >= 13 and curr_date_time.hour <=22):
					if curr_date_time.month == previous_date_time.month:
						total_of_low_tariﬀs += kwh
						counter_for_low_tariffs += 1
					else:
						tariﬀs_list.append({
							'month':month_name[previous_date_time.month],
							'average_high_tariﬀs': 0.3*(total_of_high_tariﬀs/counter_for_high_tariffs),
							'average_low_tariﬀs': 0.1*(total_of_low_tariﬀs/counter_for_low_tariffs),
							'year':previous_date_time.year
						})
						total_of_low_tariﬀs = kwh
						counter_for_low_tariffs = 1
						total_of_high_tariﬀs = 0
						counter_for_high_tariffs = 0
				# Add Last Month in the sheet
				if row == first_sheet.max_row -2:
					tariﬀs_list.append({
						'month':month_name[previous_date_time.month],
						'average_high_tariﬀs': 0.3*(total_of_high_tariﬀs/counter_for_high_tariffs),
						'average_low_tariﬀs': 0.1*(total_of_low_tariﬀs/counter_for_low_tariffs),
						'year':previous_date_time.year
					})
			# End of Calculations.

		# set calculations to the fields
		self.kw = total_of_kw/number_of_rows
		self.kwh = total_of_kwh/number_of_rows
		for item in tariffs_list:
			self.append("monthly_average_tariff", {
				"month": item['month'],
				"year": item['year'],
				"low_tariff":item['average_low_tariﬀs'],
				"high_tariff":item['average_high_tariﬀs'],
			})
		self.save()
		return True